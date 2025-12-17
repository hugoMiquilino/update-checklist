import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta


def to_excel(df):

    file_path = os.path.join("E:/", "UNIDADE D", "MEUS DOCUMENTOS", "CHECK LIST.xlsx")
    sheet_names = ["Cavalos", "Carretas"]

    def format(file_path, sheet_name):

        workbook = load_workbook(file_path)
        sheet = workbook[sheet_name]

        min_row, max_row = 1, sheet.max_row
        table = Table(displayName=sheet_name, ref=f"A{min_row}:D{max_row}")

        style = TableStyleInfo(
            name="TableStyleLight8",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True,
        )
        table.tableStyleInfo = style
        sheet.add_table(table)

        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")

        for row in range(2, max_row + 1):

            for col in range(1, 5):
                cell = sheet.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal="center")
                cell.font = Font(bold=True)

            cell = sheet[f"D{row}"]
            cell.font = Font(color="FF0000", bold=True)

            cell = sheet[f"C{row}"]
            if cell.value:
                if isinstance(cell.value, datetime):
                    cell.number_format = "DD/MMM"
                else:
                    try:
                        cell.value = datetime.strptime(
                            str(cell.value), "%Y-%m-%d %H:%M:%S"
                        )
                        cell.number_format = "DD/MMM"
                    except ValueError:
                        print(f"Formato de data inválido em C{row}, ignorando...")

        for col in range(1, sheet.max_column + 1):
            max_length = 0
            for row in range(1, max_row + 1):
                cell = sheet.cell(row=row, column=col)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[
                sheet.cell(row=1, column=col).column_letter
            ].width = (max_length + 2)

        red_fill = PatternFill(
            start_color="FFD1D7", end_color="FFD1D7", fill_type="solid"
        )
        red_font = Font(color="FF003A")
        yellow_font = Font(color="E66914")

        dxf1 = DifferentialStyle(fill=red_fill, font=red_font)
        rule1 = Rule(type="expression", dxf=dxf1, stopIfTrue=True)
        rule1.formula = ["$C2<TODAY()"]
        sheet.conditional_formatting.add(f"C2:C{max_row}", rule1)

        dxf2 = DifferentialStyle(font=red_font)
        rule2 = Rule(type="expression", dxf=dxf2, stopIfTrue=True)
        rule2.formula = ["AND($C2>=TODAY(), $C2<=TODAY()+6)"]
        sheet.conditional_formatting.add(f"C2:C{max_row}", rule2)

        dxf3 = DifferentialStyle(font=yellow_font)
        rule3 = Rule(type="expression", dxf=dxf3, stopIfTrue=True)
        rule3.formula = ["AND($C2>TODAY()+6, $C2<=TODAY()+13)"]
        sheet.conditional_formatting.add(f"C2:C{max_row}", rule3)

        rule4 = Rule(type="expression", dxf=dxf2, stopIfTrue=True)
        rule4.formula = ['OR($A2="SEM MOTORISTA", $A2="PÁTIO DE APOIO")']
        sheet.conditional_formatting.add(f"A2:A{max_row}", rule4)

        data = []
        for row in range(2, max_row + 1):
            data.append(
                [
                    sheet.cell(row=row, column=1).value,
                    sheet.cell(row=row, column=2).value,
                    sheet.cell(row=row, column=3).value,
                    sheet.cell(row=row, column=4).value,
                ]
            )

        data.sort(key=lambda x: x[2] if isinstance(x[2], datetime) else datetime.max)

        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)
        workbook.save(file_path)
        workbook.close()

        print(f"=>   Aba {sheet_name} formatada com sucesso.\n")

    for sheet_name in sheet_names:

        with pd.ExcelFile(file_path) as xls:
            existing_df = pd.read_excel(xls, sheet_name=sheet_name)

        existing_df["Vencimento"] = pd.to_datetime(
            existing_df["Vencimento"], errors="coerce", dayfirst=True
        )

        for index, row in df.iterrows():
            placa = row["Placa"]
            data = row["Vencimento"]

            existing_row = existing_df[existing_df["Placa"] == placa]
            if not existing_row.empty:
                existing_index = existing_row.index[0]

                existing_df.at[existing_index, "Vencimento"] = pd.to_datetime(
                    data, dayfirst=True, errors="coerce"
                )

        with pd.ExcelWriter(
            file_path, engine="openpyxl", mode="a", if_sheet_exists="replace"
        ) as writer:
            existing_df.to_excel(writer, sheet_name=sheet_name, index=False)

        format(file_path, sheet_name)